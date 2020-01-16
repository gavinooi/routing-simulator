import csv
import random
from datetime import datetime, timedelta
import re

import networkx as nx

from openpyxl import load_workbook, Workbook

from graph import find_path
from db_handler import DBHandler

class Simulator:

	results = []
	timeline = []
	all_orders = {}

	def __init__(self, graph_file, orders_file, output_file, cost_factor='time', algo='STATIC', clear_graph=True):
		self.dynamic = algo != 'STATIC'
		self.output_file = output_file
		print(
			'###############################\n'
			'### JANIO ROUTING SIMULATOR ###\n'
			'###############################\n'
			f'\nCONFIGURATIONS:\nGraph File:  {graph_file}\nOrder File:  {orders_file}.csv\nOutput File: {self.output_file}\n'
			f'\nADDITIONAL CONFIGURATIONS\nAlgo: {algo}\nClear graph: {clear_graph}'
		)
		self.cost_factor = cost_factor
		self.sheet_name = datetime.now().strftime("%d-%m T%H-%M-%S") + f'({cost_factor})'
		self.handler = DBHandler()
		self.orders = self._load_orders(orders_file)
		self._build_graph(graph_file, clear_graph)

	### SETUP ###

	def _load_orders(self, orders_file):
		orders_list = []
		with open(f'{orders_file}.csv') as orders:
			reader = csv.DictReader(orders)
			for row in reader:
				order_data = {}
				for key,val in row.items():
					order_data[key.lower().replace(' ', '_')] = val
				orders_list.append(order_data)
		print(f'ORDERS LOADED: {len(orders_list)}\n')
		return orders_list

	def _finish(self):
		self.handler.finish()

	def _build_graph(self, graph_file, clear_graph):
		workbook = load_workbook(f'{graph_file}.xlsx')
		node_sheet = workbook['nodes']
		nodes = []
		link_sheet = workbook['links']
		links = []

		for row in node_sheet.iter_rows(min_row=2, values_only=True):
			attr = {'name': f'"{row[0]}"', 'label': row[1]}
			if row[2]:
				attr['attr'] = f', {row[2]}'
			nodes.append(attr)

		for row in link_sheet.iter_rows(min_row=2, values_only=True):
			start = re.search("startDate:datetime\((.*?)\),endDate", row[3].replace(' ', '')).group(1)[1:-1]
			end = re.search("endDate:datetime\((.*?)\),", row[3].replace(' ', '')).group(1)[1:-1]
			for i in range(1):
				start_date = datetime.strptime(start, '%Y-%m-%dT%H:%M') + timedelta(days=i)
				end_date = datetime.strptime(end, '%Y-%m-%dT%H:%M') + timedelta(days=i)
				start_str = start_date.strftime('%Y-%m-%dT%H:%M')
				end_str = end_date.strftime('%Y-%m-%dT%H:%M')
				new = re.sub("startDate:datetime\((.*?)\),endDate", f"startDate:datetime('{start_str}'),endDate",
										 row[3].replace(' ', ''))
				new = re.sub("endDate:datetime\((.*?)\),", f"endDate:datetime('{end_str}'),", new.replace(' ', ''))

				attr = {
					'node1': f'"{row[0]}"',
					'node1_label': row[1],
					'link': row[2],
					'node2': f'"{row[4]}"',
					'node2_label': row[5],
					'attr': new + ', order_count: []'
				}
				links.append(attr)

		self.handler.build_graph(nodes, links, clear_graph)

	def _output_result(self):
		try:
			workbook = load_workbook(self.output_file)
			sheet = workbook.create_sheet(self.sheet_name)
		except:
			workbook = Workbook()
			sheet = workbook.active
			sheet.title = self.sheet_name
		finally:
			sheet.append(list(self.results[0].keys()))
			results_sorted = sorted(self.results, key=lambda i: i['tracking_no'])
			for result in results_sorted:
				sheet.append(list(result.values()))

			orders = {}
			for row in sheet.iter_rows(min_row=1,max_col=1):
				cell = row[0]
				if orders.get(cell.value):
					orders[cell.value].append(cell.coordinate)
				else:
					orders[cell.value] = [cell.coordinate]

			for k, v in orders.items():
				if len(v) > 1:
					sheet.merge_cells(f'{v[0]}:{v[-1]}')
			workbook.save(self.output_file)

	def _delay_arrival(self, initial_arrival):
		delay_weights = {
			0: 50,
			0.5: 20,
			1: 15,
			3: 10,
			5: 5
		}
		ls = []
		for delay, prob in delay_weights.items():
			ls += [delay] * prob
		delay = random.choice(ls)
		return initial_arrival + timedelta(hours=delay), delay

	def _get_actions(self, event_type):
		action_mapping = {
			'create': [self.create_order, self.increment_order_count],
			'arrive': [],
			'leave': [self.decrement_order_count],
			'expire': [self.expire_link],
			'deliver': [self.order_delivered]
		}

		if self.dynamic and event_type == 'arrive':
			return [self.insert_noise, self.reach_node]
		else:
			return action_mapping[event_type]

	### TIMELINE ###

	def add_event(self, new_event):
		'''
		adds an event into the timeline, in a sorted order
		event = {
			datetime: time the event occurs
			type: [create, arrive, leave, expire]
			desc: description of the event (with some order specific data like tracking no)
			kwargs: dictionary of keyword args to be passed into the pre stage
		}
		:return:
		'''

		if len(self.timeline) == 0:
			self.timeline.append(new_event)
		else:
			for i,event in enumerate(self.timeline):
				if new_event['datetime'] <= event['datetime']:
					self.timeline.insert(i, new_event)
					return 0

			self.timeline.append(new_event)

	def add_create_order_event(self, order):
		# add the create order event into the timeline
		tracking_no = order['tracking_no']
		created_datetime = order['created_on'][:-3] + order['created_on'][-2:]
		created_on = datetime.strptime(created_datetime, '%Y-%m-%dT%H:%M:%S.%f%z')
		origin = order['origin_zone']
		destination = order['destination_zone']
		kwargs = {
			'tracking_no': tracking_no,
			'created_on': created_on,
			'agent_app': order['agent_application_name'],
			'payment_type': order['payment_type'],
			'origin_zone': origin,
			'destination_zone': destination
		}
		event = {
			'datetime': created_on,
			'desc': f'Create order: {tracking_no}, {origin} --> {destination}',
			'actions': self._get_actions('create'),
			'kwargs': kwargs
		}

		self.add_event(event)

	def consume_event(self, event):
		'''
		runs the pre, during & post stage methods, logs event to the timeline
		:return:
		'''

		kwargs = event['kwargs']
		print(f'\nRun event: {kwargs["tracking_no"]}\n{event["desc"]}\n{str(event["datetime"])}')
		for action in event['actions']:
			kwargs = action(**kwargs)

	def run_timeline(self):
		while len(self.timeline) != 0:
			self.consume_event(self.timeline.pop(0))

	### ACTIONS ###

	def insert_noise(self, **kwargs):
		return kwargs

	def expire_link(self, **kwargs):

		self.handler.expire_link(kwargs['link'])

	def create_order(self, **kwargs):
		# load the sub graph
		tracking_no = kwargs['tracking_no']
		sub_graph = self.handler.filter_graph(kwargs)
		if not sub_graph:
			self.results.append(
				{
					'tracking_no': tracking_no,
					'cost_factor': None,
					'conditions': None,
					'path': f'No path found: {kwargs["origin_zone"]} - {kwargs["destination_zone"]}.',
					'cost': 0
				}
			)
			return {'links': [], 'tracking_no': tracking_no}

		g = nx.MultiDiGraph()
		for r in sub_graph.relationships:
			from_node = r.nodes[0]
			g.add_node(from_node['name'], label=list(from_node.labels)[0], **from_node._properties)
			to_node = r.nodes[1]
			g.add_node(to_node['name'], label=list(to_node.labels)[0], **to_node._properties)
			g.add_edge(from_node['name'], to_node['name'], attr_dict=r._properties)

		# find the path
		links, cost = find_path(g, kwargs, self.cost_factor)
		end_node = links[0][1][0]
		order_path = {
			'graph': g,
			'links': links,
			'payment_type': kwargs['payment_type'],
			'agent_app': kwargs['agent_app']
		}
		self.all_orders.update({tracking_no: order_path})

		path = f'({end_node})'
		for link in links:
			path = f'({link[0][0]}) > [{link[2]["operatedBy"]}] > ' + path

		self.results.append(
			{
				'tracking_no': tracking_no,
				'cost_factor': self.cost_factor,
				'conditions': None,
				'path': path,
				'cost': cost
			}
		)
		for link in links:
			kwargs = {
				'link': link,
				'tracking_no': tracking_no
			}
			from_node = link[0][0]
			to_node = link[1][0]
			leave_time = link[2]['startDate']

			leave_event = {
				'datetime': leave_time,
				'desc': f'Leave node: {from_node}',
				'actions': self._get_actions('leave'),
				'kwargs': kwargs
			}
			self.add_event(leave_event)

			self.add_event(
				{
					'datetime': leave_time + timedelta(minutes=30),
					'desc': f'Expire link: ({from_node}) -> ({to_node}) at {str(leave_time)}',
					'actions': self._get_actions('expire'),
					'kwargs': kwargs
				}
			)

			arrive_time = link[2]['endDate']
			delay = 0

			if self.dynamic:
				arrive_time, delay = self._delay_arrival(arrive_time)

			arrive_kwargs = {
				'tracking_no': kwargs['tracking_no'],
				'link': kwargs['link'],
				'delay': delay,
				'arrive_time': arrive_time,
			}
			arrive_event = {
				'datetime': arrive_time,
				'desc': f'Arrive node: {to_node}',
				'actions': self._get_actions('arrive'),
				'kwargs': arrive_kwargs
			}
			if to_node == end_node:
				self.add_event(
					{
						'datetime': arrive_time,
						'desc': f'Order {tracking_no} delivered to {to_node}',
						'actions': self._get_actions('deliver'),
						'kwargs': kwargs
					}
				)
			self.add_event(arrive_event)

		return {'links': links, 'tracking_no': tracking_no}

	def increment_order_count(self, **kwargs):
		self.handler.increment_order_count(**kwargs)

	def decrement_order_count(self, **kwargs):
		# remove order from a timed link
		# happens when an order is picked up or when updating order counts
		self.handler.decrement_order_count(**kwargs)

	def reach_node(self, **kwargs):
		tracking_no = kwargs['tracking_no']
		orders_dict = self.all_orders[tracking_no]
		current_node = kwargs['link'][1][0]
		end_node = orders_dict['links'][0][1][0]

		current_dict = orders_dict['graph'].nodes[current_node]
		if current_dict.get('hub_type') and current_dict['hub_type'] == 'JANIO':
			start_node = current_node
		else:
			start_node = next((link[1][0] for link in orders_dict['links'] if link[0][0] == current_node), end_node)

		if start_node == end_node:
			return {'links': [], 'tracking_no': tracking_no}

		order_details = {
			'start_label': 'WAREHOUSE',
			'origin_zone': start_node,
			'destination_zone': end_node,
			'payment_type': orders_dict['payment_type'],
			'agent_app': orders_dict['agent_app'],
			'created_on': kwargs['arrive_time']
		}
		sub_graph = self.handler.filter_graph(order_details)
		if not sub_graph:
			self.results.append(
				{
					'tracking_no': tracking_no,
					'cost_factor': None,
					'conditions': None,
					'path': f'No path found: {order_details["origin_zone"]} - {order_details["destination_zone"]}.',
					'cost': 0
				}
			)
			return {'links': [], 'tracking_no': tracking_no}

		g = nx.MultiDiGraph()
		for r in sub_graph.relationships:
			from_node = r.nodes[0]
			g.add_node(from_node['name'], label=list(from_node.labels)[0], **from_node._properties)
			to_node = r.nodes[1]
			g.add_node(to_node['name'], label=list(to_node.labels)[0], **to_node._properties)
			g.add_edge(from_node['name'], to_node['name'], attr_dict=r._properties)

		# find the path
		links, cost = find_path(g, order_details, self.cost_factor)
		path = f'({end_node})'
		for link in links:
			path = f'({link[0][0]}) > [{link[2]["operatedBy"]}] > ' + path
		self.results.append(
			{
				'tracking_no': tracking_no,
				'cost_factor': self.cost_factor,
				'conditions': f'Delay by {kwargs["delay"]}hrs',
				'path': path,
				'cost': cost
			}
		)

	def order_delivered(self, **kwargs):
		self.all_orders.pop(kwargs['tracking_no'])

	def run_simulation(self):
		print('\nBUILDING TIMELINE')
		for count, order in enumerate(self.orders):
			print(f'RUNNING ORDER {count+1}/{len(self.orders)}', end='\r')
			self.add_create_order_event(order)

		self.run_timeline()

		self._output_result()
		self._finish()
		print('\nSIMULATION FINISHED')
